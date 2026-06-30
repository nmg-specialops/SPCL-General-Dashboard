import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import load_workbook

# ======================================================
# CONFIGURATION
# ======================================================

st.set_page_config(
    page_title="SPCL General Dashboard",
    page_icon="🌍",
    layout="wide"
)

DROPBOX_URL = (
    "https://www.dropbox.com/scl/fi/"
    "pm80k4kjyzqz8yez7sffu/"
    "SPCL_DataCollection-MasterSheet_forDASHBOARD.xlsx"
    "?rlkey=3zehft3tkllkl789hdptna4f3"
    "&st=phue6k1f"
    "&dl=1"
)

# ======================================================
# LOAD WORKBOOK
# ======================================================

@st.cache_data(show_spinner=False)
def load_workbook_from_dropbox():

    response = requests.get(DROPBOX_URL)

    response.raise_for_status()

    workbook = load_workbook(
        filename=BytesIO(response.content),
        data_only=True
    )

    return workbook

# ======================================================
# LOAD DATA
# ======================================================

try:

    wb = load_workbook_from_dropbox()

    workbook_loaded = True

except Exception as e:

    workbook_loaded = False

    workbook_error = str(e)

# ======================================================
# TITLE
# ======================================================

st.title("🌍 SPCL General Dashboard")

st.caption("Live data from Dropbox Excel Workbook")

col1, col2 = st.columns([1,1])

with col1:

    if st.button("🔄 Refresh Dashboard"):

        st.cache_data.clear()

        st.rerun()

with col2:

    if workbook_loaded:

        st.success("Workbook Connected")

    else:

        st.error("Workbook Not Connected")

st.divider()

# ======================================================
# TABS
# ======================================================

summary_tab, agri_tab, production_tab, social_tab, financial_tab, other_tab = st.tabs(
    [
        "📊 Summary",
        "🌱 Agriculture",
        "🏭 Production",
        "👥 Social",
        "💰 Financial",
        "📋 Other"
    ]
)

# ======================================================
# SUMMARY
# ======================================================

with summary_tab:

    st.header("Executive Summary")

    st.info("Summary dashboard will be added after all sections are complete.")

# ======================================================
# AGRICULTURE
# ======================================================

with agri_tab:

    st.header("🌱 Agriculture")

    if workbook_loaded:

        ws = wb["Agriculture"]

        st.success("Agriculture worksheet loaded successfully.")

        st.subheader("Workbook Information")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.write("**Worksheet:** Agriculture")

        with col2:
            st.write("**Projects Found:**")
            st.write("Serendipalm Land")

        with col3:
            st.write("**Years Found:**")
            st.write("2026, 2025, 2024")

        st.divider()

        st.subheader("Preview")

        preview = []

        for row in range(5, 15):

            preview.append(
                [
                    ws.cell(row=row, column=1).value,
                    ws.cell(row=row, column=2).value,
                    ws.cell(row=row, column=3).value,
                    ws.cell(row=row, column=4).value,
                ]
            )

        df = pd.DataFrame(
            preview,
            columns=[
                "Metric",
                "Col B",
                "Col C",
                "Col D"
            ]
        )

        st.dataframe(df, use_container_width=True)

        st.divider()

        st.info(
            "Next version will automatically detect Projects, "
            "Locations, Years and populate KPI cards."
        )

    else:

        st.error(workbook_error)

# ======================================================
# PRODUCTION
# ======================================================

with production_tab:

    st.header("🏭 Production")

    st.info("Coming soon.")

# ======================================================
# SOCIAL
# ======================================================

with social_tab:

    st.header("👥 Social")

    st.info("Coming soon.")

# ======================================================
# FINANCIAL
# ======================================================

with financial_tab:

    st.header("💰 Financial")

    st.info("Coming soon.")

# ======================================================
# OTHER
# ======================================================

with other_tab:

    st.header("📋 Other")

    st.info("General worksheet will appear here.")
