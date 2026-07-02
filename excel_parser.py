"""
excel_parser.py

Parses the SPCL dashboard workbook into a structure that the
Streamlit dashboard can easily use.
"""

from openpyxl.cell.cell import MergedCell

# ------------------------------------------------------------------
# Configuration
# ------------------------------------------------------------------

STOP_HEADERS = [
    "DAF Smallholders",
    "DAF Serendipalm",
    "TOTAL DAF",
]

SMALLHOLDER_HEADERS = [
    "SPCL Smallholders",
    "Tanoobia Smallholders",
]


# ------------------------------------------------------------------
# Helper
# ------------------------------------------------------------------

def clean(value):
    if value is None:
        return ""

    return str(value).strip()

# ------------------------------------------------------------------
# Agriculture Structure
# ------------------------------------------------------------------

def agriculture_structure(sheet):

    structure = {
        "projects": {
            "Serendipalm": {
                "column": None,
                "years": [],
                "locations": {}
            },
            "SPCL Smallholders": {},
            "Tanoobia Smallholders": {},
            "All Projects": {}
        }
    }

    col = 2

    while col <= sheet.max_column:

        header = clean(sheet.cell(row=6, column=col).value)

        if header == "":
            col += 1
            continue

        years = [
            sheet.cell(row=7, column=col).value,
            sheet.cell(row=7, column=col + 1).value,
            sheet.cell(row=7, column=col + 2).value,
        ]

        # --------------------------------------------
        # Serendipalm estates
        # --------------------------------------------

        if header in [
            "Tweapease",
            "Abaam",
            "SWARF",
            "Old Cassava (other name?)",
            "Fante-Onomabo",
        ]:

            structure["projects"]["Serendipalm"]["locations"][header] = {
                "column": col,
                "years": years,
            }

        # --------------------------------------------
        # Smallholders
        # --------------------------------------------

        elif header == "SPCL Smallholders":

            structure["projects"]["SPCL Smallholders"] = {
                "column": col,
                "years": years,
            }

        elif header == "Tanoobia Smallholders":

            structure["projects"]["Tanoobia Smallholders"] = {
                "column": col,
                "years": years,
            }

        # --------------------------------------------
        # Workbook totals
        # --------------------------------------------

        elif header == "TOTAL Smallholders":

            # We'll use this later for a combined smallholder view
            structure["projects"]["Smallholders Total"] = {
                "column": col,
                "years": years,
            }

        elif header == "TOTAL Serendipalm":

            structure["projects"]["Serendipalm"]["column"] = col
            structure["projects"]["Serendipalm"]["years"] = years

        elif header == "TOTAL All Locations":

            structure["projects"]["All Projects"] = {
                "column": col,
                "years": years,
            }

            # Nothing useful after this point
            break

        col += 3

    return structure

# ------------------------------------------------------------------
# Metric Lookup
# ------------------------------------------------------------------

def get_metric(sheet, metric_name, column):

    for row in range(1, sheet.max_row + 1):

        value = clean(sheet.cell(row=row, column=1).value)

        if value == metric_name:

            return sheet.cell(row=row, column=column).value

    return None


# ------------------------------------------------------------------
# Year -> Column
# ------------------------------------------------------------------

def get_column(base_column, year):

    year_map = {
        2026: 0,
        2025: 1,
        2024: 2,
    }

    return base_column + year_map[year]

# ------------------------------------------------------------------
# List Metrics
# ------------------------------------------------------------------

def list_metrics(sheet):
    """
    Returns every metric name found in Column A.
    Useful for building the dashboard.
    """

    metrics = []

    for row in range(1, sheet.max_row + 1):

        value = clean(sheet.cell(row=row, column=1).value)

        if value != "":

            metrics.append(value)

    return metrics
