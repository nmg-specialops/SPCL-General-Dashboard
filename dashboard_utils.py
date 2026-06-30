import requests
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import streamlit as st

# =====================================================
# DROPBOX
# =====================================================

DROPBOX_URL = (
    "https://www.dropbox.com/scl/fi/"
    "pm80k4kjyzqz8yez7sffu/"
    "SPCL_DataCollection-MasterSheet_forDASHBOARD.xlsx"
    "?rlkey=3zehft3tkllkl789hdptna4f3"
    "&st=phue6k1f"
    "&dl=1"
)

# =====================================================
# LOAD WORKBOOK
# =====================================================

@st.cache_data(show_spinner=False)
def load_workbook_from_dropbox():

    response = requests.get(DROPBOX_URL)

    response.raise_for_status()

    workbook = load_workbook(
        BytesIO(response.content),
        data_only=True
    )

    return workbook


# =====================================================
# SHEET
# =====================================================

def get_sheet(workbook, sheet_name):
    return workbook[sheet_name]


# =====================================================
# HEADER HELPERS
# =====================================================

def get_header_value(sheet, row, col):
    """
    Returns the visible value of merged cells.
    """

    cell = sheet.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        return cell.value

    for merged in sheet.merged_cells.ranges:

        if cell.coordinate in merged:

            return sheet.cell(
                merged.min_row,
                merged.min_col
            ).value

    return None


# =====================================================
# PROJECTS
# =====================================================

def get_projects(sheet):

    projects = []

    for col in range(2, sheet.max_column + 1):

        value = get_header_value(sheet, 5, col)

        if value and value not in projects:

            projects.append(value)

    return projects


# =====================================================
# LOCATIONS
# =====================================================

def get_locations(sheet, project):

    locations = []

    inside = False

    for col in range(2, sheet.max_column + 1):

        project_name = get_header_value(sheet, 5, col)

        if project_name == project:

            inside = True

            location = sheet.cell(row=6, column=col).value

            if location and location not in locations:

                locations.append(location)

        elif inside:

            break

    return locations


# =====================================================
# YEARS
# =====================================================

def get_years(sheet, project, location):

    years = []

    inside = False

    for col in range(2, sheet.max_column + 1):

        project_name = get_header_value(sheet, 5, col)

        if project_name == project:

            inside = True

            loc = sheet.cell(row=6, column=col).value

            if loc == location:

                year = sheet.cell(row=7, column=col).value

                if year:
                    years.append(year)

        elif inside:

            break

    return years


# =====================================================
# FIND COLUMN
# =====================================================

def get_column(sheet, project, location, year):

    for col in range(2, sheet.max_column + 1):

        if (
            get_header_value(sheet,5,col)==project
            and sheet.cell(6,col).value==location
            and sheet.cell(7,col).value==year
        ):

            return col

    return None


# =====================================================
# FIND METRIC
# =====================================================

def get_metric(sheet, metric, project, location, year):

    column = get_column(
        sheet,
        project,
        location,
        year
    )

    if column is None:
        return None

    for row in range(1, sheet.max_row + 1):

        if sheet.cell(row=row,column=1).value == metric:

            return sheet.cell(row=row,column=column).value

    return None
